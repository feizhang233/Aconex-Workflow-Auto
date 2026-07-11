from __future__ import annotations

from collections.abc import Mapping, Set
from math import ceil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def format_table_workbook(
    path: Path,
    *,
    sheet_name: str,
    column_widths: Mapping[str, float],
    wrap_columns: Set[str] | None = None,
) -> None:
    """Apply a compact, readable table layout to a generated workbook."""
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    wrap_columns = wrap_columns or set()
    headers = {
        str(cell.value): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for header, width in column_widths.items():
        if header in headers:
            sheet.column_dimensions[_column_letter(headers[header])].width = width

    for row in sheet.iter_rows(min_row=2):
        row_height = 18
        for cell in row:
            header = str(sheet.cell(1, cell.column).value or "")
            wraps = header in wrap_columns
            cell.alignment = Alignment(vertical="top", wrap_text=wraps)
            if wraps and cell.value:
                lines = str(cell.value).splitlines() or [""]
                estimated_lines = sum(max(1, ceil(len(line) / 100)) for line in lines)
                row_height = max(row_height, min(150, 15 * estimated_lines))
        sheet.row_dimensions[row[0].row].height = row_height

    workbook.save(path)


def format_workflow_status_workbook(path: Path) -> None:
    format_table_workbook(
        path,
        sheet_name="Workflow Status",
        column_widths={
            "workflow_number": 16,
            "workflow_title": 38,
            "review_status": 26,
            "step_1_completed_time": 24,
            "step_1_due_time": 24,
            "step_1_review_status": 24,
            "step_1_overdue_duration_or_status": 18,
            "step_2_completed_time": 24,
            "step_2_due_time": 24,
            "step_2_review_status": 24,
            "step_2_overdue_duration_or_status": 18,
        },
        wrap_columns={"workflow_title"},
    )


def _column_letter(column_number: int) -> str:
    value = ""
    while column_number:
        column_number, remainder = divmod(column_number - 1, 26)
        value = chr(65 + remainder) + value
    return value
